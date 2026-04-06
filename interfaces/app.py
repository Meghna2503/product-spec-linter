import streamlit as st
import streamlit.components.v1 as components
import sys, os, json, re, math
from concurrent.futures import ThreadPoolExecutor, as_completed
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from backends import get_backend
from core.linter import PRDLinter

# ── Page config (must be first) ───────────────────────────────────────────
st.set_page_config(page_title="Product Spec Linter", page_icon="🔍", layout="wide")

# ── Theme state ───────────────────────────────────────────────────────────
if "theme" not in st.session_state:
    st.session_state.theme = "light"

# ── CSS injection — single block, theme-driven via Python variable ────────
def inject_css(theme: str):
    is_dark = theme == "dark"

    # ── Surface & background tokens ───────────────────────────────
    bg_app          = "#0f172a"   if is_dark else "#f4f6fb"
    bg_sidebar      = "#1e293b"   if is_dark else "#ffffff"
    bg_surface      = "#1e293b"   if is_dark else "#ffffff"
    bg_surface2     = "#0f172a"   if is_dark else "#f0f4fb"
    border          = "#334155"   if is_dark else "#d1d9e6"
    text_primary    = "#e2e8f0"   if is_dark else "#1a2236"
    text_muted      = "#e2e8f0"   if is_dark else "#5a6a85"
    input_bg        = "#1e293b"   if is_dark else "#ffffff"
    input_border    = "#475569"   if is_dark else "#b8c4d8"
    input_shadow    = "rgba(59,130,246,0.10)" if is_dark else "rgba(67,97,238,0.12)"
    tab_list_bg     = "#1e293b"   if is_dark else "#edf0f7"
    tab_color       = "#94a3b8"   if is_dark else "#5a6a85"
    tab_active_bg   = "#4361ee"   if is_dark else "#4361ee"
    tab_active_text = "#ffffff"   if is_dark else "#ffffff"
    fix_bg          = "rgba(16,185,129,0.1)"  if is_dark else "#f0fdf4"
    fix_border      = "rgba(16,185,129,0.3)"  if is_dark else "#bbf7d0"
    fix_text        = "#6ee7b7"   if is_dark else "#166534"
    badge_err_bg    = "rgba(239,68,68,0.2)"   if is_dark else "rgba(239,68,68,0.10)"
    badge_err_txt   = "#fca5a5"   if is_dark else "#b91c1c"
    badge_wrn_bg    = "rgba(245,158,11,0.2)"  if is_dark else "rgba(245,158,11,0.10)"
    badge_wrn_txt   = "#fcd34d"   if is_dark else "#92400e"
    badge_sug_bg    = "rgba(59,130,246,0.2)"  if is_dark else "rgba(67,97,238,0.10)"
    badge_sug_txt   = "#93c5fd"   if is_dark else "#4361ee"
    tbl_hover       = "#1e293b"   if is_dark else "#f5f7fd"
    prog_bar_bg     = "#334155"   if is_dark else "#e2e8f0"
    dim_pill_bg     = "#1e293b"   if is_dark else "#f0f4fb"

    # ── Banner shadow & tip card shadow ───────────────────────────
    banner_shadow   = "0 4px 24px rgba(30,58,138,0.18), 0 1px 4px rgba(0,0,0,0.08)" if not is_dark else "0 4px 24px rgba(0,0,0,0.40)"
    card_shadow     = "0 2px 12px rgba(67,97,238,0.10), 0 1px 3px rgba(0,0,0,0.06)" if not is_dark else "0 2px 12px rgba(0,0,0,0.30)"
    sidebar_shadow  = "2px 0 16px rgba(67,97,238,0.07)" if not is_dark else "none"
    input_focus_shadow = "0 0 0 3px rgba(67,97,238,0.18)" if not is_dark else "0 0 0 3px rgba(59,130,246,0.25)"

    st.markdown(f"""<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
html, body, [class*="css"] {{ font-family: 'Inter', sans-serif; }}

[data-testid="stAppViewContainer"] {{
    background: {bg_app};
    color: {text_primary};
}}
[data-testid="stSidebar"] {{
    background: {bg_sidebar} !important;
    border-right: 1px solid {border};
    box-shadow: {sidebar_shadow};
}}
[data-testid="stSidebar"] * {{ color: {text_primary} !important; }}
[data-testid="stHeader"] {{ background: transparent !important; }}

/* ── Sidebar inputs — model + rules — well-defined boundaries ── */
[data-testid="stSidebar"] .stSelectbox > div > div,
[data-testid="stSidebar"] .stMultiSelect > div > div {{
    background: {input_bg} !important;
    border: 1.5px solid {input_border} !important;
    border-radius: 8px !important;
    box-shadow: {card_shadow} !important;
    transition: border-color 0.2s, box-shadow 0.2s;
}}
[data-testid="stSidebar"] .stSelectbox > div > div:focus-within,
[data-testid="stSidebar"] .stMultiSelect > div > div:focus-within {{
    border-color: #4361ee !important;
    box-shadow: {input_focus_shadow} !important;
}}
[data-testid="stSidebar"] .stSelectbox label,
[data-testid="stSidebar"] .stMultiSelect label {{
    font-size: 12px !important;
    font-weight: 600 !important;
    letter-spacing: 0.04em !important;
    text-transform: uppercase !important;
    color: {text_muted} !important;
    margin-bottom: 4px !important;
}}

/* Multi-select tags */
[data-testid="stSidebar"] [data-baseweb="tag"] {{
    background: {"#1e3a5f" if is_dark else "#e8edff"} !important;
    border: 1px solid {"#3b82f6" if is_dark else "#4361ee"} !important;
    border-radius: 6px !important;
}}
[data-testid="stSidebar"] [data-baseweb="tag"] span {{
    color: {"#93c5fd" if is_dark else "#2d3fb5"} !important;
    font-size: 11px !important;
    font-weight: 600 !important;
}}

/* ── Main area inputs ── */
.stTextArea textarea {{
    background: {input_bg} !important;
    color: {text_primary} !important;
    border: 1.5px solid {input_border} !important;
    border-radius: 8px !important;
    box-shadow: {card_shadow} !important;
    transition: border-color 0.2s, box-shadow 0.2s;
}}
.stTextArea textarea:focus {{
    border-color: #4361ee !important;
    box-shadow: {input_focus_shadow} !important;
}}
.stSelectbox div[data-baseweb="select"],
.stMultiSelect div[data-baseweb="select"] {{
    background: {input_bg} !important;
    border-color: {input_border} !important;
}}

/* ── Tabs — refined, no jarring blue ── */
.stTabs [data-baseweb="tab-list"] {{
    background: {tab_list_bg} !important;
    border-radius: 10px;
    padding: 5px;
    gap: 4px;
    box-shadow: {"0 2px 8px rgba(67,97,238,0.08)" if not is_dark else "none"};
    border: 1px solid {border};
}}
.stTabs [data-baseweb="tab"] {{
    color: {tab_color} !important;
    border-radius: 7px !important;
    font-weight: 500 !important;
    font-size: 13px !important;
    padding: 6px 20px !important;
    transition: background 0.18s, color 0.18s;
    border: none !important;
    background: transparent !important;
}}
.stTabs [data-baseweb="tab"]:hover {{
    background: {"rgba(67,97,238,0.08)" if not is_dark else "rgba(59,130,246,0.10)"} !important;
    color: {"#4361ee" if not is_dark else "#93c5fd"} !important;
}}
.stTabs [aria-selected="true"] {{
    background: {tab_active_bg} !important;
    color: {tab_active_text} !important;
    border-radius: 7px !important;
    box-shadow: 0 2px 8px rgba(67,97,238,0.25) !important;
    font-weight: 600 !important;
}}
/* Remove the default Streamlit tab underline/indicator */
.stTabs [data-baseweb="tab-highlight"],
.stTabs [data-baseweb="tab-border"] {{
    display: none !important;
}}

/* ── Header / banner ── */
.header-bar {{
    background: {"linear-gradient(120deg, #1e293b 0%, #0f4c75 60%, #0c2a4a 100%)" if is_dark
                 else "linear-gradient(120deg, #1a2f6e 0%, #4361ee 60%, #3a86ff 100%)"};
    padding: 24px 32px;
    border-radius: 14px;
    margin-bottom: 24px;
    color: white;
    box-shadow: {banner_shadow};
    border: {"1px solid #334155" if is_dark else "none"};
    position: relative;
    overflow: hidden;
}}
.header-bar::after {{
    content: "";
    position: absolute;
    top: -40px; right: -40px;
    width: 200px; height: 200px;
    background: rgba(255,255,255,0.05);
    border-radius: 50%;
    pointer-events: none;
}}

/* ── Progress container ── */
.progress-container {{
    background: {bg_surface};
    border: 1px solid {border};
    border-radius: 12px;
    padding: 20px;
    margin-bottom: 16px;
    box-shadow: {card_shadow};
}}
.progress-header {{
    display: flex; align-items: center;
    justify-content: space-between; margin-bottom: 16px;
}}
.progress-title {{ font-size: 16px; font-weight: 600; color: {text_primary}; }}
.progress-stats {{ font-size: 13px; color: {text_muted}; }}
.progress-bar-bg {{
    background: {prog_bar_bg};
    border-radius: 999px; height: 8px;
    overflow: hidden; margin-bottom: 20px;
}}
.progress-bar-fill {{
    background: linear-gradient(90deg, #4361ee 0%, #3a86ff 100%);
    height: 100%; border-radius: 999px;
    transition: width 0.5s ease;
    box-shadow: 0 0 10px rgba(67,97,238,0.40);
}}
.rule-progress-item {{
    display: flex; align-items: center;
    gap: 12px; padding: 10px 0;
    border-bottom: 1px solid {border};
}}
.rule-progress-item:last-child {{ border-bottom: none; }}
.rule-icon {{
    width: 24px; height: 24px; border-radius: 50%;
    display: flex; align-items: center;
    justify-content: center; font-size: 12px;
}}
.rule-icon.running {{ background: #4361ee; animation: pulse 1.5s infinite; }}
.rule-icon.complete {{ background: #10b981; }}
.rule-icon.error {{ background: #ef4444; }}
@keyframes pulse {{
    0%, 100% {{ opacity: 1; transform: scale(1); }}
    50% {{ opacity: 0.7; transform: scale(0.95); }}
}}
@keyframes spin {{ to {{ transform: rotate(360deg); }} }}
.spinner {{
    width: 14px; height: 14px;
    border: 2px solid transparent;
    border-top-color: white; border-radius: 50%;
    animation: spin 1s linear infinite;
}}
.rule-name {{ flex: 1; font-size: 14px; color: {text_primary}; }}
.rule-status {{ font-size: 12px; color: {text_muted}; }}
.rule-status.success {{ color: #10b981; }}
.rule-status.warning {{ color: #f59e0b; }}
.rule-status.error   {{ color: #ef4444; }}
.mini-bar-bg {{
    width: 60px; height: 4px;
    background: {prog_bar_bg};
    border-radius: 2px; overflow: hidden;
}}
.mini-bar-fill {{
    height: 100%; background: #10b981;
    border-radius: 2px; transition: width 0.3s ease;
}}
.mini-bar-fill.has-issues {{ background: #f59e0b; }}
.mini-bar-fill.has-errors {{ background: #ef4444; }}

/* ── Score card ── */
.score-card {{
    background: {bg_surface};
    border: 1px solid {border};
    border-radius: 16px; padding: 28px 32px; margin-bottom: 20px;
    box-shadow: {card_shadow};
}}
.score-header {{ display: flex; align-items: center; gap: 28px; }}
.score-ring-wrap {{
    position: relative; width: 110px; height: 110px; flex-shrink: 0;
}}
.score-ring-label {{
    position: absolute; top: 50%; left: 50%;
    transform: translate(-50%, -50%);
    text-align: center; line-height: 1.1;
}}
.score-num {{ font-size: 28px; font-weight: 700; }}
.score-denom {{ font-size: 11px; color: {text_muted}; font-weight: 500; }}
.score-grade {{
    font-size: 13px; font-weight: 700;
    padding: 3px 10px; border-radius: 999px;
    display: inline-block; margin-top: 4px;
}}
.score-right {{ flex: 1; }}
.score-title {{ font-size: 20px; font-weight: 700; margin-bottom: 4px; color: {text_primary}; }}
.score-summary {{ font-size: 14px; margin-bottom: 16px; line-height: 1.5; color: {text_muted}; }}
.score-dims {{ display: flex; flex-wrap: wrap; gap: 8px; }}
.dim-pill {{
    display: flex; align-items: center; gap: 6px;
    background: {dim_pill_bg}; border: 1px solid {border};
    color: {text_primary};
    border-radius: 8px; padding: 6px 12px; font-size: 12px;
}}
.dim-bar-wrap {{ width: 72px; height: 6px; background: {prog_bar_bg}; border-radius: 999px; overflow: hidden; }}
.dim-bar {{ height: 100%; border-radius: 999px; }}
.dim-label {{ font-weight: 500; }}
.score-verdict {{
    margin-top: 16px; padding: 10px 14px;
    border-radius: 8px; font-size: 13px;
    font-weight: 500; line-height: 1.5;
}}

/* ── Metric boxes ── */
.metric-box {{
    background: {bg_surface};
    border: 1px solid {border};
    border-radius: 10px; padding: 16px 12px; text-align: center;
    box-shadow: {card_shadow};
}}
.metric-num   {{ font-size: 32px; font-weight: 700; line-height: 1; }}
.metric-label {{ font-size: 12px; color: {text_muted}; margin-top: 4px; font-weight: 500; }}
.metric-error   .metric-num {{ color: #ef4444; }}
.metric-warning .metric-num {{ color: #f59e0b; }}
.metric-sug     .metric-num {{ color: {"#4361ee" if not is_dark else "#3b82f6"}; }}

/* ── Severity badges ── */
.sev-badge {{
    display: inline-block; font-size: 10px; font-weight: 700;
    text-transform: uppercase; letter-spacing: .06em;
    padding: 2px 8px; border-radius: 999px; white-space: nowrap;
}}
.badge-error   {{ background: {badge_err_bg}; color: {badge_err_txt}; border: 1px solid rgba(239,68,68,0.3); }}
.badge-warning {{ background: {badge_wrn_bg}; color: {badge_wrn_txt}; border: 1px solid rgba(245,158,11,0.3); }}
.badge-sug     {{ background: {badge_sug_bg}; color: {badge_sug_txt}; border: 1px solid rgba(67,97,238,0.30); }}

/* ── Findings table ── */
.findings-table {{
    width: 100%; border-collapse: collapse;
    font-size: 13px; margin-top: 12px;
}}
.findings-table th {{
    background: {bg_surface2};
    color: {text_muted};
    font-weight: 600; font-size: 11px;
    text-transform: uppercase; letter-spacing: .05em;
    padding: 10px 12px; text-align: left;
    white-space: nowrap;
    border-bottom: 2px solid {border};
}}
.findings-table td {{
    padding: 12px 12px; vertical-align: top;
    line-height: 1.5; border-bottom: 1px solid {border};
    color: {text_primary};
}}
.findings-table tr:hover td {{ background: {tbl_hover}; }}
.col-num      {{ width: 32px; text-align: center; font-size: 11px; }}
.col-sev      {{ width: 100px; }}
.col-category {{ width: 160px; font-weight: 500; color: {text_primary}; }}
.col-issue    {{ width: 28%; color: {text_muted}; }}
.col-fix      {{ width: 35%; }}
.fix-code {{
    background: {fix_bg};
    border: 1px solid {fix_border};
    color: {fix_text};
    border-radius: 6px; padding: 8px 12px;
    font-family: inherit; font-size: 12.5px;
    white-space: pre-wrap; word-break: break-word;
    display: block; line-height: 1.5;
}}


/* ── ALL input labels — force visible in both modes ── */
label,
[data-testid="stWidgetLabel"] p,
[data-testid="stWidgetLabel"] span,
.stTextInput label,
.stTextInput [data-testid="stWidgetLabel"] p,
.stSelectbox label,
.stSelectbox [data-testid="stWidgetLabel"] p,
.stMultiSelect label,
.stMultiSelect [data-testid="stWidgetLabel"] p,
.stTextArea label,
.stTextArea [data-testid="stWidgetLabel"] p,
.stRadio label,
.stRadio [data-testid="stWidgetLabel"] p,
.stRadio > div > div > label > div > p,
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] [data-testid="stWidgetLabel"] p {{
    color: {text_primary} !important;
    opacity: 1 !important;
}}

/* ── Radio button option text — force full brightness ── */
.stRadio > div > label,
.stRadio > div > label p,
.stRadio > div > label span,
.stRadio [data-testid="stMarkdownContainer"] p,
div[data-testid="stRadio"] label,
div[data-testid="stRadio"] label p,
div[data-testid="stRadio"] label div p {{
    color: {text_primary} !important;
    opacity: 1 !important;
}}

/* ── All widget labels in main content (Tab panels) ── */
.stTabs [data-testid="stWidgetLabel"] p,
.stTabs [data-testid="stWidgetLabel"] span,
.stTabs label,
.stTabs .stRadio label,
.stTabs .stTextInput label,
.stTabs .stTextArea label,
.stTabs .stSelectbox label {{
    color: {text_primary} !important;
    opacity: 1 !important;
}}

/* ── Dark mode: placeholder text — bump contrast ── */
[data-testid="stSidebar"] .stSelectbox input::placeholder,
[data-testid="stSidebar"] .stMultiSelect input::placeholder,
.stTextInput input::placeholder,
.stTextArea textarea::placeholder,
input::placeholder {{
    color: {"#64748b" if is_dark else "#a0aec0"} !important;
    opacity: 1 !important;
}}

/* ── Input selected value text in dark ── */
[data-testid="stSidebar"] .stSelectbox [data-baseweb="select"] [data-testid="stMarkdownContainer"] p,
[data-testid="stSidebar"] .stSelectbox input,
[data-testid="stSidebar"] [data-baseweb="select"] span {{
    color: {text_primary} !important;
}}

/* ── Clear credentials / secondary action buttons ── */
button[kind="secondary"],
.stButton > button:not([kind="primary"]) {{
    background: {"#1e3a52" if is_dark else "#f4f6fb"} !important;
    color: {"#e2e8f0" if is_dark else "#2d3fb5"} !important;
    border: 1.5px solid {"#3d5068" if is_dark else "#c5cfe8"} !important;
    border-radius: 8px !important;
}}
.stButton > button:not([kind="primary"]):hover {{
    background: {"#264d6d" if is_dark else "#e8edff"} !important;
    border-color: {"#4361ee" if is_dark else "#4361ee"} !important;
}}

/* ── Top info/tip banner bar (Jira tip) ── */
[data-testid="stInfo"],
div[data-testid="stAlert"] {{
    background: {"#1a3347" if is_dark else "#eef4ff"} !important;
    border: 1px solid {"#2d5070" if is_dark else "#c7d9f8"} !important;
    color: {"#93c5fd" if is_dark else "#2d3fb5"} !important;
    border-radius: 10px !important;
    box-shadow: {"0 2px 12px rgba(0,0,0,0.30)" if is_dark else "0 2px 8px rgba(67,97,238,0.08)"} !important;
}}
[data-testid="stInfo"] p,
div[data-testid="stAlert"] p,
div[data-testid="stAlert"] div {{
    color: {"#93c5fd" if is_dark else "#2d3fb5"} !important;
}}
/* ── Export buttons ── */
.stDownloadButton > button {{
    font-size: 13px !important; padding: 8px 20px !important;
    height: 36px !important; min-height: 36px !important;
    border-radius: 6px !important; white-space: nowrap !important;
}}

/* ── Sidebar tip / caption cards ── */
[data-testid="stSidebar"] .stCaption {{
    background: {"#253347" if is_dark else "#f0f4fb"} !important;
    border: 1px solid {"#3d5068" if is_dark else "#d1d9e6"} !important;
    border-radius: 8px !important;
    padding: 10px 12px !important;
    margin-bottom: 8px !important;
    font-size: 12px !important;
    color: {"#e2e8f0" if is_dark else "#5a6a85"} !important;
    box-shadow: {"0 2px 12px rgba(0,0,0,0.30)" if is_dark else "0 2px 12px rgba(67,97,238,0.10), 0 1px 3px rgba(0,0,0,0.06)"} !important;
    display: block !important;
}}

/* ── Jira story cards ── */
.story-card {{
    background: {bg_surface};
    border: 1px solid {border};
    border-radius: 8px; padding: 14px; margin-bottom: 8px;
    box-shadow: {card_shadow};
}}
.story-key   {{ font-size: 12px; font-weight: 700; color: #4361ee; margin-bottom: 4px; }}
.story-title {{ font-size: 14px; font-weight: 600; color: {text_primary}; }}
.story-meta  {{ font-size: 11px; color: {text_muted}; margin-top: 4px; }}
</style>""", unsafe_allow_html=True)


# ── Apply CSS on every run ────────────────────────────────────────────────
inject_css(st.session_state.get("theme", "dark"))

# ── Constants ─────────────────────────────────────────────────────────────
RULES = ["AMBIGUITY","MISSING_AC","CONTRADICTION","DEPENDENCY_GAP","COMPLETENESS","EDGE_CASES"]
RULE_LABELS = {
    "AMBIGUITY":      "Ambiguous Language",
    "MISSING_AC":     "Missing Acceptance Criteria",
    "CONTRADICTION":  "Contradiction",
    "DEPENDENCY_GAP": "Dependency Gap",
    "COMPLETENESS":   "Completeness",
    "EDGE_CASES":     "Edge Case Detection",
}
SEV_CONFIG = {
    "ERROR":      ("🔴", "badge-error"),
    "WARNING":    ("🟡", "badge-warning"),
    "SUGGESTION": ("⚪", "badge-sug"),
}
DIM_CONFIG = {
    "MISSING_AC":     ("Missing Acceptance Criteria", 25, 10),
    "AMBIGUITY":      ("Ambiguous Language",          20,  8),
    "COMPLETENESS":   ("Completeness",                20,  6),
    "CONTRADICTION":  ("Contradiction",               20, 12),
    "DEPENDENCY_GAP": ("Dependency Gap",              10,  8),
    "EDGE_CASES":     ("Edge Case Detection",         15,  6),
}
GRADE_CONFIG = [
    (90, "A", "#166534", "#dcfce7"),
    (75, "B", "#1e40af", "#dbeafe"),
    (60, "C", "#92400e", "#fef3c7"),
    (40, "D", "#c2410c", "#ffedd5"),
    (0,  "F", "#991b1b", "#fee2e2"),
]

# ── Scoring ───────────────────────────────────────────────────────────────
def compute_score(all_findings, selected_rules):
    """
    Scoring: each finding deducts a % of that dimension's weight.
      ERROR      → 20% per finding
      WARNING    → 10% per finding
      SUGGESTION →  3% per finding
    Total deduction capped at 100% per dimension.
    This prevents a single ERROR from zeroing out an entire dimension.
    """
    dim_scores = {}
    PCT_RATES = {"ERROR": 0.20, "WARNING": 0.10, "SUGGESTION": 0.03}
    for rule in selected_rules:
        if rule not in DIM_CONFIG: continue
        label, weight, _ = DIM_CONFIG[rule]
        rule_findings = [f for f in all_findings if f.get("rule") == rule]
        total_pct = sum(PCT_RATES.get(f.get("severity", "SUGGESTION"), 0.03) for f in rule_findings)
        total_pct = min(total_pct, 1.0)
        dim_score = round(weight * (1 - total_pct))
        dim_scores[rule] = {"label": label, "weight": weight, "score": dim_score,
                            "pct": round((dim_score / weight) * 100)}
    total_weight = sum(v["weight"] for v in dim_scores.values())
    if total_weight == 0: return 100, dim_scores
    overall = round((sum(v["score"] for v in dim_scores.values()) / total_weight) * 100)
    return overall, dim_scores

def grade_from_score(score):
    for threshold, letter, text_col, bg_col in GRADE_CONFIG:
        if score >= threshold: return letter, text_col, bg_col
    return "F", "#991b1b", "#fee2e2"

def verdict_text(score, dim_scores):
    weak_dims = [v["label"] for v in dim_scores.values() if v["pct"] < 60]
    if score >= 90:
        return "✅ Excellent spec — ready for development with minimal risk.", "#166534", "#dcfce7"
    elif score >= 75:
        missing = f" Focus on: {', '.join(weak_dims)}." if weak_dims else ""
        return f"🟡 Solid spec with room to improve.{missing}", "#92400e", "#fef3c7"
    elif score >= 60:
        missing = f" Key gaps: {', '.join(weak_dims)}." if weak_dims else ""
        return f"🟠 Needs work before handing to engineers.{missing}", "#c2410c", "#ffedd5"
    else:
        missing = f" Critical gaps in: {', '.join(weak_dims)}." if weak_dims else ""
        return f"🔴 High risk — this spec will cause rework if shipped as-is.{missing}", "#991b1b", "#fee2e2"

# ── Render helpers ────────────────────────────────────────────────────────
def render_score_card(score, dim_scores):
    grade, grade_color, grade_bg = grade_from_score(score)
    verdict, verdict_color, verdict_bg = verdict_text(score, dim_scores)
    radius, cx, cy, stroke_w = 44, 55, 55, 10
    circumference = 2 * math.pi * radius
    filled = circumference * (score / 100)

    pills_html = ""
    for v in dim_scores.values():
        pct = v["pct"]
        bar_color = "#22c55e" if pct >= 75 else "#f59e0b" if pct >= 50 else "#ef4444"
        pills_html += (
            f'''<div class="dim-pill">'''
            f'''<div class="dim-label">{v["label"]}</div>'''
            f'''<div class="dim-bar-wrap"><div class="dim-bar" style="width:{pct}%;background:{bar_color}"></div></div>'''
            f'''<span style="font-size:11px;font-weight:600">{pct}%</span>'''
            f'''</div>'''
        )

    st.markdown(f"""
    <div class="score-card">
        <div class="score-header">
            <div class="score-ring-wrap">
                <svg width="110" height="110" viewBox="0 0 110 110">
                    <circle cx="{cx}" cy="{cy}" r="{radius}"
                        fill="none" stroke="#334155" stroke-width="{stroke_w}"/>
                    <circle cx="{cx}" cy="{cy}" r="{radius}"
                        fill="none" stroke="{grade_color}" stroke-width="{stroke_w}"
                        stroke-dasharray="{filled:.1f} {circumference:.1f}"
                        stroke-linecap="round"
                        transform="rotate(-90 {cx} {cy})"/>
                </svg>
                <div class="score-ring-label">
                    <div class="score-num" style="color:{grade_color}">{score}</div>
                    <div class="score-denom">/100</div>
                    <div class="score-grade" style="color:{grade_color};background:{grade_bg}">Grade {grade}</div>
                </div>
            </div>
            <div class="score-right">
                <div class="score-title">Spec Quality Score</div>
                <div class="score-summary">Based on {len(dim_scores)} dimensions checked by the linter.</div>
                <div class="score-dims">{pills_html}</div>
                <div class="score-verdict" style="color:{verdict_color};background:{verdict_bg}">
                    {verdict}
                </div>
            </div>
        </div>
    </div>""", unsafe_allow_html=True)

def render_metrics(summary):
    cols = st.columns(3)
    for col, (key, icon, label, cls) in zip(cols, [
        ("ERROR",      "🔴", "Errors",      "metric-error"),
        ("WARNING",    "🟡", "Warnings",    "metric-warning"),
        ("SUGGESTION", "⚪", "Suggestions", "metric-sug"),
    ]):
        with col:
            st.markdown(f"""<div class="metric-box {cls}">
                <div class="metric-num">{summary[key]}</div>
                <div class="metric-label">{icon} {label}</div>
            </div>""", unsafe_allow_html=True)

def render_findings_table(all_findings):
    rows = ""
    for i, f in enumerate(all_findings, 1):
        sev = f.get("severity", "SUGGESTION")
        icon, badge_cls = SEV_CONFIG.get(sev, SEV_CONFIG["SUGGESTION"])
        category   = RULE_LABELS.get(f.get("rule", ""), f.get("rule", ""))
        issue      = f.get("issue", "")
        suggestion = f.get("suggestion", "")
        conflict   = f.get("conflict_with", "")
        conflict_html = (f'''<br><span style="color:#7c3aed;font-size:11px">''' +
                         f'''↔ conflicts with: {conflict}</span>''') if conflict else ""
        rows += f"""<tr>
            <td class="col-num">{i}</td>
            <td class="col-sev"><span class="sev-badge {badge_cls}">{icon} {sev}</span></td>
            <td class="col-category">{category}{conflict_html}</td>
            <td class="col-issue">{issue}</td>
            <td class="col-fix"><span class="fix-code">{suggestion}</span></td>
        </tr>"""
    st.markdown(f"""
    <table class="findings-table">
        <thead><tr>
            <th class="col-num">#</th>
            <th class="col-sev">Severity</th>
            <th class="col-category">Category</th>
            <th class="col-issue">What's Wrong</th>
            <th class="col-fix">✅ Fix — select &amp; copy into Jira</th>
        </tr></thead>
        <tbody>{rows}</tbody>
    </table>""", unsafe_allow_html=True)

def build_excel(all_findings, score, dim_scores):
    import io
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None
    wb = Workbook()
    HDR  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    FILL_H = PatternFill("solid", fgColor="1e3a5f")
    grade, gc, gb = grade_from_score(score)

    ws_score = wb.active; ws_score.title = "Score Summary"
    ws_score.column_dimensions["A"].width = 30
    ws_score.column_dimensions["B"].width = 20
    ws_score.cell(1, 1, "Spec Quality Score").font = Font(name="Calibri", bold=True, size=14)
    ws_score.cell(2, 1, "Overall Score").font = Font(name="Calibri", bold=True)
    ws_score.cell(2, 2, f"{score}/100")
    ws_score.cell(3, 1, "Grade").font = Font(name="Calibri", bold=True)
    ws_score.cell(3, 2, grade)
    ws_score.cell(5, 1, "Dimension").font = HDR; ws_score.cell(5, 1).fill = FILL_H
    ws_score.cell(5, 2, "Score %").font = HDR;   ws_score.cell(5, 2).fill = FILL_H
    for i, v in enumerate(dim_scores.values(), 6):
        ws_score.cell(i, 1, v["label"]); ws_score.cell(i, 2, f"{v['pct']}%")

    ws = wb.create_sheet("Findings")
    fills = {
        "ERROR":      PatternFill("solid", fgColor="fee2e2"),
        "WARNING":    PatternFill("solid", fgColor="fef3c7"),
        "SUGGESTION": PatternFill("solid", fgColor="dbeafe"),
    }
    headers = ["#", "Severity", "Category", "What's Wrong", "Suggested Fix", "Status"]
    widths  = [4, 12, 24, 50, 60, 12]
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HDR; cell.fill = FILL_H
        cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 22
    for i, f in enumerate(all_findings, 1):
        sev  = f.get("severity", "SUGGESTION")
        fill = fills.get(sev, fills["SUGGESTION"])
        vals = [i, sev, RULE_LABELS.get(f.get("rule", ""), f.get("rule", "")),
                f.get("issue", ""), f.get("suggestion", ""), "☐ To Do"]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=i+1, column=c, value=v)
            cell.font = Font(name="Calibri", size=11); cell.fill = fill
            cell.alignment = Alignment(
                horizontal="left" if c > 2 else "center",
                vertical="top", wrap_text=True)
        ws.row_dimensions[i+1].height = 60
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()

# ── Main run function ─────────────────────────────────────────────────────

def _display_results(all_findings, summary, score, dim_scores, selected_rules):
    """Re-render results from cache without re-running the model."""
    grade_letter = grade_from_score(score)[0]
    report_out = {
        "score": score, "grade": grade_letter,
        "dimensions": {k: v["pct"] for k, v in dim_scores.items()},
        "summary": summary, "total": len(all_findings), "findings": all_findings,
    }
    md_lines = [f"# Lint Report — Score {score}/100 (Grade {grade_letter})\n"] + [
        f"## [{f.get('severity','')}] {RULE_LABELS.get(f.get('rule',''),f.get('rule',''))}\n\n"
        f"**What's wrong:** {f.get('issue','')}\n\n"
        f"**Fix:** {f.get('suggestion','')}\n"
        for f in all_findings
    ]
    excel_bytes = build_excel(all_findings, score, dim_scores)
    b1, b2, b3, _ = st.columns([1, 1, 1, 4])
    with b1:
        st.download_button("📊 Export Excel",
            data=excel_bytes if excel_bytes else b"", file_name="lint-report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, disabled=excel_bytes is None)
    with b2:
        st.download_button("📝 Markdown",
            data="\n".join(md_lines), file_name="lint-report.md",
            mime="text/markdown", use_container_width=True)
    with b3:
        st.download_button("🔧 JSON",
            data=json.dumps(report_out, indent=2), file_name="lint-report.json",
            mime="application/json", use_container_width=True)
    st.divider()
    render_score_card(score, dim_scores)
    render_metrics(summary)
    if not all_findings:
        st.success("🎉 No issues found — this is a clean spec!", icon="✅")
        return

    total = len(all_findings)

    # ── Filter bar ────────────────────────────────────────────────
    st.markdown("<p style='margin:16px 0 6px;font-size:13px'>"
                f"<strong>{total} issue{'s' if total > 1 else ''} found</strong> "
                "— filter below, copy fixes directly into Jira</p>",
                unsafe_allow_html=True)

    fc1, fc2, fc3 = st.columns([1.2, 1.8, 0.6])
    with fc1:
        sev_filter = st.multiselect(
            "🔍 Severity",
            options=["ERROR", "WARNING", "SUGGESTION"],
            default=["ERROR", "WARNING", "SUGGESTION"],
            key=f"filter_sev_{id(all_findings)}",
            label_visibility="collapsed",
            placeholder="Filter by severity…"
        )
    with fc2:
        available_cats = sorted({RULE_LABELS.get(f.get("rule",""), f.get("rule","")) for f in all_findings})
        cat_filter = st.multiselect(
            "🗂 Category",
            options=available_cats,
            default=available_cats,
            key=f"filter_cat_{id(all_findings)}",
            label_visibility="collapsed",
            placeholder="Filter by category…"
        )
    with fc3:
        reset_key = f"reset_filter_{id(all_findings)}"
        if st.button("✕ Reset", key=reset_key, use_container_width=True):
            for k in list(st.session_state.keys()):
                if k.startswith(f"filter_sev_{id(all_findings)}") or k.startswith(f"filter_cat_{id(all_findings)}"):
                    del st.session_state[k]
            st.rerun()

    filtered = [
        f for f in all_findings
        if f.get("severity", "SUGGESTION") in sev_filter
        and RULE_LABELS.get(f.get("rule",""), f.get("rule","")) in cat_filter
    ]

    shown = len(filtered)
    if shown < total:
        st.caption(f"Showing {shown} of {total} issues")

    st.divider()
    if not filtered:
        st.info("No issues match the current filters.", icon="🔍")
    else:
        render_findings_table(filtered)

def run_and_show(spec_text, model, selected_rules, progress_area=None, source="tab1"):
    backend = get_backend("ollama", model=model)
    linter  = PRDLinter(backend)
    st.session_state["_progress_html"] = ""
    if progress_area is None:
        progress_area = st.empty()
    # Pre-build fixed layout: header + 6 individual slots, 3 per column
    _prog_header = st.empty()
    _prog_col1, _prog_col2 = st.columns(2)
    _rule_slots = [_prog_col1.empty(), _prog_col1.empty(), _prog_col1.empty(),
                   _prog_col2.empty(), _prog_col2.empty(), _prog_col2.empty()]

    rule_info = {
        r: {"status": "running", "findings": [], "errors": 0, "warnings": 0, "suggestions": 0}
        for r in selected_rules
    }
    all_findings   = []
    summary        = {"ERROR": 0, "WARNING": 0, "SUGGESTION": 0}
    completed_count = [0]

    def render_progress(done=False):
        total     = len(selected_rules)
        completed = completed_count[0]

        if done:
            header = f"✅ **All checks complete** — {completed}/{total} rules analysed"
        else:
            header = f"⏳ **Analysing with {model}…** — {completed}/{total} complete"

        rule_lines = []
        for r in selected_rules:
            info   = rule_info[r]
            status = info["status"]
            label  = RULE_LABELS[r]
            if status == "running":
                icon   = "🔄"
                detail = "Analysing…"
            elif status == "complete":
                if info["errors"] > 0:
                    icon   = "🔴"
                    detail = f"{info['errors']} error(s), {info['warnings']} warning(s)"
                elif info["warnings"] > 0:
                    icon   = "🟡"
                    detail = f"{info['warnings']} warning(s)"
                else:
                    icon   = "✅"
                    detail = "No issues"
            else:
                icon   = "❌"
                detail = "Failed"
            rule_lines.append(f"{icon} **{label}** — {detail}")

        # Update header
        _prog_header.markdown(f"##### {header}")

        # Pad rule_lines to always fill 6 slots (in case fewer rules selected)
        padded = rule_lines + [""] * (6 - len(rule_lines))
        for slot, text in zip(_rule_slots, padded):
            slot.markdown(text)

    render_progress()


    with ThreadPoolExecutor(max_workers=len(selected_rules)) as executor:
        future_to_rule = {
            executor.submit(linter._run_rule, spec_text, rule): rule
            for rule in selected_rules
        }
        for future in as_completed(future_to_rule):
            rule = future_to_rule[future]
            try:
                findings = future.result()
                all_findings.extend(findings)
                for f in findings:
                    sev = f.get("severity", "SUGGESTION")
                    if sev in summary: summary[sev] += 1
                errors   = sum(1 for f in findings if f.get("severity") == "ERROR")
                warnings = sum(1 for f in findings if f.get("severity") == "WARNING")
                rule_info[rule].update({
                    "status": "complete", "findings": findings,
                    "errors": errors, "warnings": warnings,
                    "suggestions": len(findings) - errors - warnings,
                })
            except Exception:
                rule_info[rule]["status"] = "error"
                rule_info[rule]["errors"] = 1
            completed_count[0] += 1
            render_progress()

    render_progress(done=True)
    st.divider()

    score, dim_scores = compute_score(all_findings, selected_rules)

    # Store results under source-specific keys so tab1 and tab2 never overwrite each other
    st.session_state[f"lint_findings_{source}"] = all_findings
    st.session_state[f"lint_summary_{source}"]  = summary
    st.session_state[f"lint_score_{source}"]    = score
    st.session_state[f"lint_dim_scores_{source}"] = dim_scores

    # ── Export bar — TOP, before findings ─────────────────────────────────
    grade_letter = grade_from_score(score)[0]
    report_out = {
        "score": score, "grade": grade_letter,
        "dimensions": {k: v["pct"] for k, v in dim_scores.items()},
        "summary": summary, "total": len(all_findings), "findings": all_findings,
    }
    md_lines = [f"# Lint Report — Score {score}/100 (Grade {grade_letter})\n"] + [
        f"## [{f.get('severity','')}] {RULE_LABELS.get(f.get('rule',''),f.get('rule',''))}\n\n"
        f"**What's wrong:** {f.get('issue','')}\n\n"
        f"**Fix:** {f.get('suggestion','')}\n"
        for f in all_findings
    ]
    excel_bytes = build_excel(all_findings, score, dim_scores)

    b1, b2, b3, _ = st.columns([1, 1, 1, 4])
    with b1:
        st.download_button("📊 Export Excel",
            data=excel_bytes if excel_bytes else b"", file_name="lint-report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, disabled=excel_bytes is None)
    with b2:
        st.download_button("📝 Markdown",
            data="\n".join(md_lines), file_name="lint-report.md",
            mime="text/markdown", use_container_width=True)
    with b3:
        st.download_button("🔧 JSON",
            data=json.dumps(report_out, indent=2), file_name="lint-report.json",
            mime="application/json", use_container_width=True)

    st.divider()
    render_score_card(score, dim_scores)
    render_metrics(summary)

    if not all_findings:
        st.success("🎉 No issues found — this is a clean spec!", icon="✅")
        return

    total = len(all_findings)

    # ── Filter bar ────────────────────────────────────────────────
    st.markdown("<p style='margin:16px 0 6px;font-size:13px'>"
                f"<strong>{total} issue{'s' if total > 1 else ''} found</strong> "
                "— filter below, copy fixes directly into Jira</p>",
                unsafe_allow_html=True)

    fc1, fc2, fc3 = st.columns([1.2, 1.8, 0.6])
    with fc1:
        sev_filter = st.multiselect(
            "🔍 Severity",
            options=["ERROR", "WARNING", "SUGGESTION"],
            default=["ERROR", "WARNING", "SUGGESTION"],
            key=f"filter_sev_{id(all_findings)}",
            label_visibility="collapsed",
            placeholder="Filter by severity…"
        )
    with fc2:
        available_cats = sorted({RULE_LABELS.get(f.get("rule",""), f.get("rule","")) for f in all_findings})
        cat_filter = st.multiselect(
            "🗂 Category",
            options=available_cats,
            default=available_cats,
            key=f"filter_cat_{id(all_findings)}",
            label_visibility="collapsed",
            placeholder="Filter by category…"
        )
    with fc3:
        reset_key = f"reset_filter_{id(all_findings)}"
        if st.button("✕ Reset", key=reset_key, use_container_width=True):
            for k in list(st.session_state.keys()):
                if k.startswith(f"filter_sev_{id(all_findings)}") or k.startswith(f"filter_cat_{id(all_findings)}"):
                    del st.session_state[k]
            st.rerun()

    filtered = [
        f for f in all_findings
        if f.get("severity", "SUGGESTION") in sev_filter
        and RULE_LABELS.get(f.get("rule",""), f.get("rule","")) in cat_filter
    ]

    shown = len(filtered)
    if shown < total:
        st.caption(f"Showing {shown} of {total} issues")

    st.divider()
    if not filtered:
        st.info("No issues match the current filters.", icon="🔍")
    else:
        render_findings_table(filtered)

# ── Header ────────────────────────────────────────────────────────────────
st.markdown("""
<div class="header-bar">
    <h2 style="margin:0;font-size:22px">🔍 Product Spec Linter</h2>
    <p style="margin:6px 0 0;opacity:.8;font-size:14px">
        Catch ambiguity, contradictions, and missing ACs — before they reach your dev team.<br>
        <span style="opacity:.6;font-size:12px">🔒 Runs locally via Ollama. Your spec never leaves your machine.</span>
    </p>
</div>""", unsafe_allow_html=True)

# ── Sidebar ───────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ Options")
    new_dark = st.toggle("Dark Mode", value=st.session_state.theme == "dark")
    if new_dark != (st.session_state.theme == "dark"):
        st.session_state.theme = "dark" if new_dark else "light"
        st.rerun()
    st.divider()
    model          = st.selectbox("Model", [
        # ── Cloud frontier ──────────────────────
        "kimi-k2.5:cloud",
        "claude-opus-4-5",
        "claude-sonnet-4-5",
        "claude-haiku-3-5",
        "gpt-4o",
        "gpt-4o-mini",
        "gpt-4-turbo",
        "gemini-2.5-pro",
        "gemini-2.0-flash",
        "gemini-1.5-pro",
        # ── Local (Ollama) ───────────────────────
        "llama3",
        "mistral",
        "phi3:mini",
        "deepseek-r1:7b",
        "qwen2.5:7b",
    ])
    selected_rules = st.multiselect("Rules", RULES, default=RULES)
    st.divider()
    st.caption("🔒 Ollama models run 100% locally.\nYour spec never leaves your machine.")
    st.caption("💡 **Tip:** Run one rule first to test before the full suite.")

# ── Tabs ──────────────────────────────────────────────────────────────────
tab1, tab2 = st.tabs(["📝 Paste Text", "🔗 Connect Jira"])

# ════════════════════════════════════════════════════════
# TAB 1 — Paste spec
# ════════════════════════════════════════════════════════
with tab1:
    st.subheader("📄 Your Spec")
    spec_text = st.text_area("", height=180, key="paste_input",
        placeholder=(
            "Paste your user story, epic, or requirement text here...\n\n"
            "Works with Notion, Linear, Shortcut, GitHub Issues, or plain text."
        ))
    if spec_text:
        st.caption(f"📝 {len(spec_text.split())} words · {len(spec_text)} characters")

    # ── Dynamic ETA ───────────────────────────────────────────────
    # ── Per-model ETA (seconds per rule, parallel run) ───────────
    MODEL_ETA = {
        # Cloud frontier — network + inference latency
        "kimi-k2.5:cloud":   45,
        "claude-opus-4-5":   50,
        "claude-sonnet-4-5": 35,
        "claude-haiku-3-5":  20,
        "gpt-4o":            35,
        "gpt-4o-mini":       18,
        "gpt-4-turbo":       45,
        "gemini-2.5-pro":    40,
        "gemini-2.0-flash":  18,
        "gemini-1.5-pro":    35,
        # Local (Ollama) — CPU/GPU dependent
        "llama3":            35,
        "mistral":           30,
        "phi3:mini":         20,
        "deepseek-r1:7b":    40,
        "qwen2.5:7b":        32,
    }
    # Fallback: categorise unknown models so ETA is never just a shrug
    def _fallback_eta(m: str) -> int:
        m = m.lower()
        if any(x in m for x in ["opus", "gpt-4", "pro", "ultra", "large"]):
            return 50   # large cloud frontier
        if any(x in m for x in ["sonnet", "flash", "turbo", "4o", "haiku"]):
            return 30   # fast cloud
        if any(x in m for x in ["mini", "nano", "tiny", "small", "phi"]):
            return 18   # small / fast local
        if any(x in m for x in ["cloud", "api", "online"]):
            return 40   # generic cloud
        return 35       # safe default (local medium)
    secs_per_rule  = MODEL_ETA.get(model, _fallback_eta(model))
    # Rules run in parallel — ETA ≈ slowest rule (1 rule worth of time) + small overhead per extra rule
    eta_secs       = secs_per_rule + max(0, len(selected_rules) - 1) * 5
    eta_min        = eta_secs // 60
    eta_sec        = eta_secs % 60
    if eta_min >= 1:
        eta_str = f"~{eta_min} min {eta_sec}s" if eta_sec else f"~{eta_min} min"
    else:
        eta_str = f"~{eta_sec}s"
    num_rules_str  = f"{len(selected_rules)} rule{'s' if len(selected_rules) != 1 else ''}"
    st.caption(f"⏱️ Estimated time: **{eta_str}** · {num_rules_str} running in parallel")
    run = st.button("▶ Run Lint", type="primary", key="run_paste")
    if run:
        if not spec_text.strip():
            st.warning("Please paste some spec text first.", icon="⚠️")
        elif not selected_rules:
            st.warning("Select at least one rule in the sidebar.", icon="⚠️")
        else:
            word_count = len(spec_text.strip().split())
            if word_count < 15:
                st.error(
                    f"**❌ Not enough content to lint** — your input is only "
                    f"**{word_count} word{'s' if word_count != 1 else ''}**. "
                    "A valid spec needs a role, an action, a goal, and at least "
                    "one acceptance criterion (15+ words).",
                    icon="🚫"
                )
                st.markdown("""**Example of a minimal valid spec:**
```
As a registered user, I want to reset my password via email
so that I can regain access to my account.

Acceptance Criteria:
  Given I am on the login page
  When I click "Forgot Password" and enter my email
  Then I receive a reset link within 60 seconds
```""")
                st.stop()
            st.divider()
            st.subheader("📋 Lint Report")
            _prog_area = st.empty()
            run_and_show(spec_text, model, selected_rules, _prog_area, source="tab1")

    elif "lint_findings_tab1" in st.session_state and st.session_state["lint_findings_tab1"] is not None:
        st.divider()
        st.subheader("📋 Lint Report")
        _display_results(
            st.session_state["lint_findings_tab1"],
            st.session_state["lint_summary_tab1"],
            st.session_state["lint_score_tab1"],
            st.session_state["lint_dim_scores_tab1"],
            selected_rules
        )

# ════════════════════════════════════════════════════════
# TAB 2 — Jira
# ════════════════════════════════════════════════════════
with tab2:
    for key in ["jira_url", "jira_email", "jira_token"]:
        if key not in st.session_state: st.session_state[key] = ""

    st.info("Connect Jira to fetch a story + its full epic — then lint for cross-story issues.", icon="💡")
    col1, col2 = st.columns([1, 2], gap="large")

    with col1:
        st.subheader("🔗 Jira Settings")
        jira_url   = st.text_input("Jira URL",   key="jira_url",   placeholder="https://yourcompany.atlassian.net")
        jira_email = st.text_input("Email",       key="jira_email", placeholder="you@company.com")
        jira_token = st.text_input("API Token",   key="jira_token", type="password",
                        help="Create at: https://id.atlassian.com/manage-profile/security/api-tokens")
        if st.button("🗑️ Clear credentials", use_container_width=True):
            st.session_state["jira_url"] = ""
            st.session_state["jira_email"] = ""
            st.session_state["jira_token"] = ""
            st.rerun()
        st.divider()
        issue_key_raw = st.text_input("Story ID",
            placeholder="MDP-3  or  https://company.atlassian.net/browse/MDP-3")
        match     = re.search(r'([A-Z][A-Z0-9_]+-\d+)', issue_key_raw.strip())
        issue_key = match.group(1) if match else issue_key_raw.strip()
        if issue_key and issue_key != issue_key_raw.strip():
            st.caption(f"✅ Extracted key: **{issue_key}**")
        scope    = st.radio("Check against", ["Entire epic (cross-story)", "This story only"], horizontal=True)
        # ── Dynamic ETA (Jira) ────────────────────────────────────────
        secs_per_rule_j = MODEL_ETA.get(model, _fallback_eta(model))
        eta_secs_j      = secs_per_rule_j + max(0, len(selected_rules) - 1) * 5
        eta_min_j       = eta_secs_j // 60
        eta_sec_j       = eta_secs_j % 60
        if eta_min_j >= 1:
            eta_str_j = f"~{eta_min_j} min {eta_sec_j}s" if eta_sec_j else f"~{eta_min_j} min"
        else:
            eta_str_j = f"~{eta_sec_j}s"
        st.caption(f"⏱️ Estimated time: **{eta_str_j}** · {len(selected_rules)} rule(s) in parallel")
        run_jira = st.button("▶ Fetch & Lint", type="primary", use_container_width=True, key="run_jira")
        st.markdown("""<div style="background:#f0f9ff;border:1px solid #bae6fd;border-radius:8px;
            padding:12px;margin-top:16px;font-size:12px;color:#0369a1;">
            🔒 Credentials used this session only, never stored.<br>
            All AI runs locally via Ollama.<br>
            <a href="https://id.atlassian.com/manage-profile/security/api-tokens" target="_blank">Create a read-only token →</a>
        </div>""", unsafe_allow_html=True)

    with col2:
        if not run_jira:
            st.markdown("""<div style="background:#f8fafc;border:2px dashed #cbd5e1;border-radius:10px;
                padding:60px;text-align:center;color:#64748b;margin-top:32px">
                <div style="font-size:40px;margin-bottom:12px">🔗</div>
                <div style="font-weight:600;margin-bottom:6px">Connect Jira and enter a Story ID</div>
                <div style="font-size:13px">Fetches story + full epic, lints for cross-story issues</div>
            </div>""", unsafe_allow_html=True)
        elif not all([jira_url, jira_email, jira_token, issue_key]):
            st.warning("Please fill in all Jira fields.", icon="⚠️")
        elif not selected_rules:
            st.warning("Select at least one rule in the sidebar.", icon="⚠️")
        else:
            jira_spec_text = None
            with st.spinner(f"Connecting to Jira and fetching {issue_key}..."):
                try:
                    from connectors.jira_connector import JiraConnector
                    jira = JiraConnector(jira_url, jira_email, jira_token)
                    if "epic" in scope.lower():
                        stories        = jira.get_epic_stories(issue_key)
                        jira_spec_text = jira.format_for_linter(stories, focus_key=issue_key)
                        st.success(f"✅ Fetched **{len(stories)} stories** from epic.")
                        with st.expander(f"📋 View fetched stories ({len(stories)})"):
                            for s in stories:
                                focus = s["key"] == issue_key
                                st.markdown(
                                    f'''<div class="story-card" style="{'border-color:#0f4c75;border-width:2px;' if focus else ''}">'''
                                    f'''<div class="story-key">{s["key"]} {'⭐ FOCUS' if focus else ''}</div>'''
                                    f'''<div class="story-title">{s["summary"]}</div>'''
                                    f'''<div class="story-meta">Status: {s["status"]} · Type: {s["type"]}</div>'''
                                    f'''</div>''',
                                    unsafe_allow_html=True)
                    else:
                        story          = jira.get_story(issue_key)
                        jira_spec_text = jira.format_for_linter([story])
                        st.success(f"✅ Fetched story **{issue_key}**")
                except Exception as e:
                    st.error(f"Jira connection failed: {e}", icon="❌")
                    if any(x in str(e) for x in ["401", "credential", "token", "Unauthorized"]):
                        st.info("💡 Double-check your Jira URL, email, and API token.")
            if jira_spec_text:
                st.divider()
                _prog_area2 = st.empty()
                run_and_show(jira_spec_text, model, selected_rules, _prog_area2, source="tab2")

            elif "lint_findings_tab2" in st.session_state and st.session_state["lint_findings_tab2"] is not None:
                st.divider()
            st.subheader("📋 Lint Report")
            _display_results(
                st.session_state["lint_findings_tab2"],
                st.session_state["lint_summary_tab2"],
                st.session_state["lint_score_tab2"],
                st.session_state["lint_dim_scores_tab2"],
                selected_rules
            )
